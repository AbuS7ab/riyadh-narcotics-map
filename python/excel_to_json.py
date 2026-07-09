from openpyxl import load_workbook
import json

# اسم ملف الإكسل
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent

excel_file = BASE_DIR / "القطاع الشمالي احداثيات.xlsx"
json_file = BASE_DIR / "data" / "facilities.json"

wb = load_workbook(excel_file)
ws = wb.active

facilities = []

# الصف الثاني هو رؤوس الأعمدة
for row in ws.iter_rows(min_row=3, values_only=True):

    if not row[1]:
        continue

    coords = str(row[8]).split(",")

    if len(coords) != 2:
        continue

    lat = float(coords[0].strip())
    lng = float(coords[1].strip())

    facilities.append({

        "name": row[1],
        "type": row[2],
        "license": str(row[3]),
        "district": row[6],
        "street": row[7],
        "sector": row[10],
        "google_maps": row[9],
        "lat": lat,
        "lng": lng

    })

with open(json_file, "w", encoding="utf-8") as f:
    json.dump(facilities, f, ensure_ascii=False, indent=4)

print(f"تم إنشاء {json_file}")
print(f"عدد المنشآت: {len(facilities)}")